import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# 样式定义
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
title_font = Font(name='Arial', bold=True, size=14)
subheader_font = Font(name='Arial', bold=True, size=10)
data_font = Font(name='Arial', size=10)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(cell, is_subheader=False):
    cell.font = header_font if not is_subheader else subheader_font
    cell.fill = header_fill if not is_subheader else PatternFill(start_color='D6DCE5', end_color='D6DCE5', fill_type='solid')
    cell.alignment = center_align
    cell.border = thin_border

def style_data(cell):
    cell.font = data_font
    cell.alignment = left_align
    cell.border = thin_border

# ==================== Sheet 1: 五行方位花种总表 ====================
ws1 = wb.active
ws1.title = '五行-方位-花种总表'

# 标题
ws1.merge_cells('A1:G1')
ws1['A1'] = '五行·方位·花种·颜色 核心规则表'
ws1['A1'].font = title_font
ws1['A1'].alignment = center_align

headers = ['方位', '五行', '代表色', '相生花种', '推荐颜色', '宜忌说明', '吉祥寓意']
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=2, column=col, value=h)
    style_header(c)

data_rows = [
    # 东方
    ['东', '木', '青、绿', '绿萝、吊兰、文竹、虎皮兰、发财树', '绿色、青色', '木气旺盛，宜青绿色植物，可促进健康和事业运', '木代表肝，东方旺则身体健康、事业上进'],
    ['东南', '木', '青、绿', '绿萝、吊兰、龟背竹、常春藤', '绿色、粉绿', '木气略弱于正东，适合爬藤类，助人际关系', '利于文昌运、贵人运'],
    # 南方
    ['南', '火', '红、紫', '红掌、鸿运当头、一品红、扶桑花', '红色、紫色', '火气旺，宜红花但要适量，避免过度炽热', '利名声、利学业、利社交'],
    ['西南', '土', '黄、棕', '君子兰、虎皮兰、黄金葛', '黄色、棕色', '土气主稳定，宜中型绿植，平衡坤位', '利人际关系、婚姻、财运'],
    # 西方
    ['西', '金', '白、金', '白掌、茉莉花、多肉植物、仙人掌', '白色、金色', '金气肃杀，宜白色花或硬质植物，忌红色', '利财运、利武职、利白事'],
    ['西北', '金', '白、金', '白掌、栀子花、玉兰花、虎尾兰', '白色、银色', '金气充沛，宜白色花助事业运', '利贵人运、利决策力'],
    # 北方
    ['北', '水', '蓝、黑', '水仙、风信子、蓝色妖姬、睡莲', '蓝色、黑色', '水气旺，宜水养或蓝紫色花，忌黄色', '利事业、利智谋、利收藏'],
    ['东北', '土', '黄、咖啡', '腊梅、黄玫瑰、迎春花、杜鹃', '黄色、橙色', '土气旺，宜耐寒花木，忌绿色过盛', '利稳中求进、防小人'],
    # 中央
    ['中', '土', '黄、紫', '君子兰、兰花、紫罗兰、万年青', '黄色、紫色', '土居中调和，宜稳重花卉，忌过于张扬', '利整体运势、调和五行'],
]

for row_idx, row_data in enumerate(data_rows, 3):
    for col_idx, val in enumerate(row_data, 1):
        c = ws1.cell(row=row_idx, column=col_idx, value=val)
        style_data(c)

for col in 'ABCDEFG':
    ws1.column_dimensions[col].width = 18

# ==================== Sheet 2: 房间功能推荐 ====================
ws2 = wb.create_sheet('房间功能推荐')

ws2.merge_cells('A1:F1')
ws2['A1'] = '房间功能·花卉摆放·风水宜忌'
ws2['A1'].font = title_font
ws2['A1'].alignment = center_align

headers2 = ['房间', '功能定位', '推荐花种', '推荐颜色', '摆放位置', '风水宜忌']
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=2, column=col, value=h)
    style_header(c)

room_data = [
    ['客厅', '招财旺宅、待客', '幸福树、平安树、发财树、鸿运当头', '绿色、红色', '电视柜旁、角落、主位对面', '宜：大门对角线（财位）\n忌：正对大门、阻挡通道'],
    ['卧室', '助眠安神、夫妻和睦', '茉莉花、薰衣草、吊兰、君子兰', '白色、淡紫色', '床头柜、窗台、梳妆台旁', '宜：小型、淡香\n忌：大叶、过香、剧毒'],
    ['书房', '旺文昌、聚精会神', '文竹、绿萝、菖蒲、铜钱草', '绿色、青色', '书桌、书架、文昌位', '宜：小巧精致、有书卷气\n忌：带刺、过于艳丽'],
    ['厨房', '去油净气、添活力', '薄荷、罗勒、常春藤、吊兰', '绿色、白色', '窗台、橱柜上方、冰箱顶', '宜：香草、可食用\n忌：花粉过多、招蚊虫'],
    ['卫生间', '去晦气、转运势', '常春藤、吊兰、虎皮兰、吸毒草', '绿色、白色', '洗手台旁、角落、窗台', '宜：水养、耐阴\n忌：枯萎、凋谢'],
    ['阳台', '接气纳财、旺宅', '月季、绣球、三角梅、铁树', '红、粉、黄', '阳光充足处、角落', '宜：向阳、易养\n忌：枯枝败叶'],
    ['玄关', '迎气送气、第一印象', '绿萝、虎皮兰、发财树、富贵竹', '绿色、金色', '鞋柜上方、入门两侧', '宜：高大、常绿\n忌：带刺、正对门'],
    ['餐厅', '增进食欲、和睦', '小型月季、蝴蝶兰、黄丽', '粉色、黄色', '餐桌旁、餐边柜', '宜：小型、清香\n忌：过大、枯萎'],
]

for row_idx, row_data in enumerate(room_data, 3):
    for col_idx, val in enumerate(row_data, 1):
        c = ws2.cell(row=row_idx, column=col_idx, value=val)
        style_data(c)

for col in 'ABCDEF':
    ws2.column_dimensions[col].width = 20

# ==================== Sheet 3: 花种属性库 ====================
ws3 = wb.create_sheet('花种属性库')

ws3.merge_cells('A1:J1')
ws3['A1'] = '花卉植物属性大全（五行/颜色/花语/功效）'
ws3['A1'].font = title_font
ws3['A1'].alignment = center_align

headers3 = ['花种', '五行', '颜色', '花语', '风水功效', '毒性', '难度', '花期', '光照', '适合位置']
for col, h in enumerate(headers3, 1):
    c = ws3.cell(row=2, column=col, value=h)
    style_header(c)

flower_data = [
    ['绿萝', '木', '绿', '坚韧、善良', '旺财、净化空气', '无', '★', '全年', '耐阴', '客厅/书房/卫生间'],
    ['吊兰', '木', '绿/白', '无奈、无奈的爱', '去煞气、旺宅', '无', '★', '全年', '耐阴', '客厅/阳台/卧室'],
    ['文竹', '木', '绿', '永恒、友情', '旺文昌、利学业', '无', '★★★', '全年', '散光', '书房/客厅'],
    ['虎皮兰', '金', '绿/黄', '刚毅、坚强', '旺财、去煞气', '无', '★', '全年', '耐阴', '客厅/卧室/玄关'],
    ['发财树', '木', '绿', '招财进宝、事业有成', '旺财、旺事业', '无', '★★', '5-11月', '散光', '客厅/办公室'],
    ['幸福树', '木', '绿', '幸福美满', '家庭和睦、旺宅', '无', '★★', '5-9月', '散光', '客厅/阳台'],
    ['平安树', '木', '绿', '平安吉祥', '保平安、护健康', '无', '★★', '冬季', '散光', '客厅/玄关'],
    ['鸿运当头', '火', '红', '好运连连、完美', '招好运、旺家运', '无', '★★', '全年', '散光', '客厅/卧室'],
    ['红掌', '火', '红', '热情、大展宏图', '招财、旺人缘', '微毒', '★★', '全年', '散光', '客厅/办公室'],
    ['白掌', '金', '白', '纯洁、和平', '事业顺利、一帆风顺', '无', '★', '5-8月', '散光', '客厅/书房'],
    ['君子兰', '火/土', '红/黄', '高贵、君子之风', '旺家运、利人际关系', '无', '★★★', '冬春', '散光', '客厅/书房'],
    ['蝴蝶兰', '木/火', '粉/紫', '幸福飞来、爱你', '旺姻缘、增感情', '无', '★★★★', '2-5月', '散光', '客厅/卧室'],
    ['茉莉花', '金', '白', '纯洁、忠贞', '旺感情、添香气', '无', '★★★', '5-8月', '全阳', '阳台/庭院'],
    ['栀子花', '金', '白', '坚强、永恒的爱', '旺桃花、旺学业', '无', '★★★', '5-7月', '全阳', '阳台/庭院'],
    ['薄荷', '木', '绿', '真诚、温暖', '去晦气、清醒头脑', '无', '★', '5-10月', '全阳', '厨房/阳台'],
    ['铜钱草', '水', '绿', '财运滚滚、花开富贵', '旺财、水生财', '无', '★', '全年', '散光', '客厅/书房'],
    ['水仙', '水', '白/黄', '自恋、思念', '旺文才、增智慧', '有毒', '★★', '冬季', '全阳', '客厅/书房'],
    ['风信子', '水', '蓝/紫', '胜利、竞技', '旺事业、聚人气', '有毒', '★★', '3-4月', '全阳', '阳台/窗台'],
    ['仙人掌', '金', '绿', '坚强、奇迹', '防辐射、去煞气', '无', '★', '夏季', '全阳', '阳台/书桌'],
    ['多肉植物', '金', '多彩', '坚韧、可爱', '旺土气、防小人', '无', '★', '全年', '全阳', '阳台/书桌'],
    ['月季', '火', '红/粉', '爱情、美丽', '旺姻缘、添魅力', '无', '★★', '4-9月', '全阳', '阳台/庭院'],
    ['绣球', '火', '蓝/粉', '希望、团聚', '旺人缘、助姻缘', '无', '★★', '5-7月', '散光', '阳台/庭院'],
    ['三角梅', '火', '红/紫', '热情、坚韧', '旺财运、增活力', '无', '★', '4-11月', '全阳', '阳台/庭院'],
    ['常春藤', '木', '绿', '忠诚、永恒', '去煞气、保平安', '无', '★', '全年', '耐阴', '客厅/阳台'],
    ['龟背竹', '木', '绿', '健康、长寿', '旺健康、增福寿', '无', '★★', '全年', '散光', '客厅/卧室'],
    ['薰衣草', '火', '紫', '等待、浪漫', '助睡眠、旺姻缘', '无', '★★★', '5-8月', '全阳', '卧室/阳台'],
    ['黄金葛', '木', '绿/黄', '财运亨通、吉祥', '旺财、水生金', '无', '★', '全年', '耐阴', '客厅/书房'],
    ['万年青', '土', '绿', '健康长寿、吉祥', '旺宅、保平安', '有毒', '★', '全年', '耐阴', '客厅/玄关'],
    ['富贵竹', '木', '绿', '富贵吉祥、竹报平安', '旺财、旺学业', '无', '★', '全年', '耐阴', '客厅/书房'],
]

for row_idx, row_data in enumerate(flower_data, 3):
    for col_idx, val in enumerate(row_data, 1):
        c = ws3.cell(row=row_idx, column=col_idx, value=val)
        style_data(c)

for col_idx in range(1, 11):
    ws3.column_dimensions[get_column_letter(col_idx)].width = 14

# ==================== Sheet 4: 摆放位置宜忌 ====================
ws4 = wb.create_sheet('摆放位置宜忌')

ws4.merge_cells('A1:E1')
ws4['A1'] = '花卉摆放位置详细指南'
ws4['A1'].font = title_font
ws4['A1'].alignment = center_align

headers4 = ['位置', '宜', '忌', '原因', '推荐植物']
for col, h in enumerate(headers4, 1):
    c = ws4.cell(row=2, column=col, value=h)
    style_header(c)

placement_data = [
    ['大门正对', '高大绿植、常绿植物', '带刺植物、空盆', '大门是气口，宜导气入宅，不宜阻挡或泄气', '发财树、幸福树、富贵竹'],
    ['大门对角线（财位）', '花期长、颜色鲜艳', '枯萎、凋谢', '财位宜生气旺盛，忌萧条', '鸿运当头、红掌、蝴蝶兰'],
    ['客厅主位（沙发背后）', '高大、稳重', '藤蔓、爬墙', '主位宜有靠山，藤蔓主攀附、阴气', '虎皮兰、平安树、幸福树'],
    ['电视柜两侧', '常绿、净化空气', '过高遮挡视线', '对称摆放有利气场平衡', '绿萝、吊兰、龟背竹'],
    ['窗台/阳台', '向阳、易养、花期长', '枯枝败叶、过敏花卉', '阳光充足利于光合作用，枯萎招阴', '月季、三角梅、绣球'],
    ['卧室床头', '小型、淡香、无毒', '过香、剧毒、大叶', '睡眠时需清净气场，过香影响睡眠', '茉莉、薰衣草、吊兰'],
    ['书房书桌', '文气、小巧精致', '带刺、过于艳丽', '书房需安静文气，不宜冲煞', '文竹、铜钱草、小型绿萝'],
    ['厨房窗台', '香草、可食用', '花粉过多、招虫', '厨房油烟多，香草可净化', '薄荷、罗勒、迷迭香'],
    ['卫生间角落', '耐阴、水养', '喜阳、干燥', '卫生间阴气重，需耐阴去阴', '常春藤、吊兰、虎皮兰'],
    ['楼梯下方/角落', '常绿、向上生长', '枯萎、藤蔓下垂', '角落宜生气凝聚，藤蔓下垂主下坡', '绿萝、黄金葛、常春藤'],
    ['横梁下方', '小型、圆形植物', '高大、正对', '横梁压顶主压力，小株缓冲', '小型多肉、仙人掌'],
    ['镜子对面', '小型、矮壮', '高大、藤蔓', '镜子反射，植物太高易形成"惊"象', '小型绿植、多肉'],
    ['神位前', '清雅、淡色', '艳丽、大红大紫', '神位前宜清净，艳丽主喧宾夺主', '小白掌、淡色兰花'],
]

for row_idx, row_data in enumerate(placement_data, 3):
    for col_idx, val in enumerate(row_data, 1):
        c = ws4.cell(row=row_idx, column=col_idx, value=val)
        style_data(c)

for col in 'ABCDE':
    ws4.column_dimensions[col].width = 22

# ==================== Sheet 5: 季节推荐 ====================
ws5 = wb.create_sheet('季节推荐')

ws5.merge_cells('A1:F1')
ws5['A1'] = '四季花卉推荐与养护指南'
ws5['A1'].font = title_font
ws5['A1'].alignment = center_align

headers5 = ['季节', '五行', '旺位', '推荐花种', '养护要点', '风水布局重点']
for col, h in enumerate(headers5, 1):
    c = ws5.cell(row=2, column=col, value=h)
    style_header(c)

season_data = [
    ['春（2-4月）', '木', '东方、东南', '杜鹃、迎春花、桃花、君子兰、风信子', '浇水适量、增加光照、换盆施肥', '木气旺盛，宜青绿色，正东、东南旺事业人缘'],
    ['夏（5-7月）', '火', '南方', '茉莉、栀子花、三角梅、绣球、米兰', '遮阴防晒、增加浇水、注意通风', '火气过旺，宜白色、淡色花，中和南方火'],
    ['秋（8-10月）', '金', '西方、西北', '菊花、桂花、白掌、一品红、蟹爪兰', '减少浇水、增加磷钾肥、修剪枯枝', '金气旺，宜白色花，西方助财运'],
    ['冬（11-1月）', '水', '北方', '水仙、风信子、君子兰、鸿运当头、腊梅', '防寒保暖、控制浇水、增加光照', '水气旺，宜红色、紫色，北方利事业'],
]

for row_idx, row_data in enumerate(season_data, 3):
    for col_idx, val in enumerate(row_data, 1):
        c = ws5.cell(row=row_idx, column=col_idx, value=val)
        style_data(c)

for col in 'ABCDEF':
    ws5.column_dimensions[col].width = 20

# ==================== Sheet 6: 使用说明 ====================
ws6 = wb.create_sheet('使用说明')

ws6.merge_cells('A1:B1')
ws6['A1'] = '风水花卉推荐系统 使用指南'
ws6['A1'].font = title_font

guide_content = [
    ['一、核心逻辑', ''],
    ['1. 输入条件', '房屋朝向 + 房间位置 + （可选）主人五行'],
    ['2. 匹配规则', '方位 → 五行 → 颜色 → 花种'],
    ['3. 输出内容', '推荐花种 + 颜色 + 摆放位置 + 风水解释'],
    ['', ''],
    ['二、快速查询步骤', ''],
    ['Step 1', '查看「五行-方位-花种总表」，确定方位对应的五行和颜色'],
    ['Step 2', '查看「房间功能推荐」，确定该房间的功能定位'],
    ['Step 3', '在「花种属性库」中筛选符合五行+颜色的花种'],
    ['Step 4', '查看「摆放位置宜忌」，确认具体位置注意事项'],
    ['Step 5', '参考「季节推荐」，选择当季适合的花卉'],
    ['', ''],
    ['三、优先级原则', ''],
    ['第一优先级', '房间功能（卧室不能用有毒/过香）'],
    ['第二优先级', '方位五行（金位不用绿色）'],
    ['第三优先级', '个人喜好（在满足前两条前提下）'],
    ['', ''],
    ['四、常见组合推荐', ''],
    ['东向客厅', '发财树/幸福树 + 绿色 + 主位旁'],
    ['南向卧室', '茉莉/薰衣草 + 淡紫色 + 床头柜'],
    ['西向书房', '白掌/虎皮兰 + 白色 + 书桌旁'],
    ['北向客厅', '鸿运当头/红掌 + 红色 + 财位'],
    ['卫生间', '常春藤/吊兰 + 绿色 + 角落'],
    ['厨房', '薄荷/罗勒 + 绿色 + 窗台'],
    ['', ''],
    ['五、知识库维护', ''],
    ['更新频率', '每季度更新一次花种数据'],
    ['数据来源', '传统风水文献 + 现代植物学'],
    ['注意事项', '风水学有一定主观性，本表仅供参考'],
]

for row_idx, (col_a, col_b) in enumerate(guide_content, 2):
    c_a = ws6.cell(row=row_idx, column=1, value=col_a)
    c_b = ws6.cell(row=row_idx, column=2, value=col_b)
    if col_a and col_b:
        c_a.font = Font(name='Arial', bold=True, size=10)
        c_b.font = Font(name='Arial', size=10)
    elif col_a:
        c_a.font = Font(name='Arial', bold=True, size=11)

ws6.column_dimensions['A'].width = 20
ws6.column_dimensions['B'].width = 50

# 保存
output_path = '/Users/sger/WorkBuddy/20260411082948/风水花卉推荐知识库.xlsx'
wb.save(output_path)
print(f'已保存到: {output_path}')
